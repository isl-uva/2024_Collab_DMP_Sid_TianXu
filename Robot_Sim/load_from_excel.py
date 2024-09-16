import pandas as pd
import numpy as np
from openpyxl import load_workbook
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

def load_demo_library_xlsx(filename):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    # Read the data and convert it to a DataFrame
    data = sheet.values
    df = pd.DataFrame(data)

    # Split by blank rows
    df = df[1:len(df)]
    # df_list = np.split(df, df[df.isnull().all(1)].index)
    indices = np.where(df.isnull().all(axis=1))[0]
    df_list = np.split(df, indices)
    

    demo_waypoints = np.empty(len(df_list), dtype=object)
    demo_orientations = np.empty(len(df_list), dtype=object)

    # Process dataframes
    for i in range(1, len(df_list) + 1):
        df_list[i - 1] = df_list[i - 1].dropna(how='all')
        array = df_list[i-1].astype(float).to_numpy()
        sliced_waypoints = array[:, :3]
        sliced_orientations = array[:, 3:]

        waypoints = []
        orientations = []
        for j in range(len(sliced_waypoints)):
            waypoints.append(sliced_waypoints[j].tolist())

        for j in range(len(sliced_orientations)):
            orientations.append(sliced_orientations[j].tolist())

        demo_waypoints[i-1] = waypoints
        demo_orientations[i-1] = orientations

    return demo_waypoints, demo_orientations



def load_task_configuration(filename):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    # Read the data and convert it to a DataFrame
    data = sheet.values
    df = pd.DataFrame(data)

    # excludes the row title
    df = df[1:len(df)]

    task_waypoints = np.empty(1, dtype=object)
    task_orientations = np.empty(1, dtype=object)
    
    array = df.astype(float).to_numpy()
    
    waypoints = []
    orientations = []

    for i in range(len(array)):
         waypoints.append(array[i,:3].tolist())
         orientations.append(array[i,3:].tolist())
         
    task_waypoints = waypoints
    task_orientations = orientations 

    return task_waypoints, task_orientations

